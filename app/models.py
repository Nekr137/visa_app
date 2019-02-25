from django.db import models
import datetime
import openpyxl,re,os,datetime
from django.http import HttpResponseRedirect,HttpResponse, FileResponse
from openpyxl.writer.excel import save_virtual_workbook
from django.forms import ModelChoiceField

ulr = openpyxl.styles.Border(
    top=openpyxl.styles.Side(style='thin'),
    left=openpyxl.styles.Side(style='thin'),
    right=openpyxl.styles.Side(style='thin'),
)

blr = openpyxl.styles.Border(
    bottom=openpyxl.styles.Side(style='thin'),
    left=openpyxl.styles.Side(style='thin'),
    right=openpyxl.styles.Side(style='thin'),
)

br = openpyxl.styles.Border(
    bottom=openpyxl.styles.Side(style='thin'),
    right=openpyxl.styles.Side(style='thin'),
)

b = openpyxl.styles.Border(
    bottom=openpyxl.styles.Side(style='thin'),
)

t = openpyxl.styles.Border(
    top=openpyxl.styles.Side(style='thin'),
)

l = openpyxl.styles.Border(
    left=openpyxl.styles.Side(style='thin'),
)

tlrb = openpyxl.styles.Border(
    top=openpyxl.styles.Side(style='thin'),
    left=openpyxl.styles.Side(style='thin'),
    right=openpyxl.styles.Side(style='thin'),
    bottom=openpyxl.styles.Side(style='thin')
)


def date_format(d):
    """Дата в формате dd.mm.yyyy"""
    d = re.findall(r'\d+', str(d))
    d.reverse()
    return '/'.join(d)

def SplitText(t,lengths):
    """
    Делит строку на список строк заданной длины. lengths = [10,6,7,...]
    :param t:   текстовая строка
    :param lengths:   список длин строк
    :return:    список строк
    """
    # сформировали список пустых строк
    strings = []
    for k in range(len(lengths)):   strings.append('')

    # индекс для букав и индекс для массива длин строк
    i = l = 0

    for w in t.split():              # для каждого слова
        i+=len(w)
        if i>lengths[l]:            # не вышли ли за пределы длины строки
            l+=1
            i = 0
        strings[l] += w + ' '
    return strings


class Partners(models.Model):
    partner = models.TextField()
    default = models.BooleanField(default=False)
    def __str__(self):
        return self.partner

class VisaNumber(models.Model):
    visanumber = models.IntegerField()
    def __str__(self):
        return str(self.visanumber)

class AdditionalInfo(models.Model):
    info = models.TextField()
    default = models.BooleanField(default=False)
    def __str__(self):
        return self.info

class Placements(models.Model):
    placement = models.TextField()
    default = models.BooleanField(default=False)
    def __str__(self):
        return self.placement

class Organizations(models.Model):
    organization = models.TextField()
    default = models.BooleanField(default=False)
    def __str__(self):
        return self.organization

class Nationality(models.Model):
    nationality = models.CharField(max_length=100)
    default = models.BooleanField(default=False)
    def __str__(self):
        return self.nationality


class Routs(models.Model):
    rout = models.TextField()
    default = models.BooleanField(default=False)
    def __str__(self):
        return self.rout

class Ships(models.Model):
    name = models.CharField(max_length=30)
    default = models.BooleanField(default=False)
    def __str__(self):
        return self.name


class Dates(models.Model):
    ship = models.ForeignKey(Ships, on_delete=models.CASCADE)
    entry = models.DateField()
    departure = models.DateField()
    default = models.BooleanField(default=False)
    def __str__(self):
        return str(self.entry) + '/' + str(self.departure)



class Form2(models.Model):
    firstname = models.CharField(max_length=30)
    familyname = models.CharField(max_length=30)
    lastname = models.CharField(max_length=30)
    name = models.CharField(max_length=30)
    sex = models.CharField(max_length=10)
    goal = models.CharField(max_length=30)
    birthday = models.DateField()
    passport = models.TextField()
    multiplicity = models.TextField()
    partner = models.TextField()
    invitation_number = models.TextField()
    nationality = models.TextField()
    entry = models.DateField()
    departure = models.DateField()
    placement = models.TextField()
    rout = models.TextField()
    hostorganization = models.TextField()
    additionalinfo = models.TextField()
    date = models.DateField()  # document date (hidden)
    visa_type = models.TextField()  # type of visa (single, group). Hidden

    def __str__(self):
        return 'Familyname: ' + self.familyname + ' Name: ' + self.name


    def FormXlsx_group(self,fin):
        self.fname = fin
        self.wb = openpyxl.Workbook()
        self.wb = openpyxl.load_workbook(filename=self.fname)

        #sheet1 = self.wb["Лист1"]
        sheet1 = self.wb.active
        if str(self.partner).lower() == 'genvisa':
            sheet1['F2'] = '130319 -                USA ' + \
            datetime.datetime.strptime(str(self.entry), '%Y-%m-%d').date().strftime("%d/%m")

        sheet1['B6'] = 'визовое приглашение № ' + str(self.invitation_number)
        sheet1['K6'] = 'визовое приглашение № ' + str(self.invitation_number)
        sheet1['D8'] = self.multiplicity
        sheet1['D10'] = self.nationality
        sheet1['C12'], sheet1['L12'] = [date_format(self.entry)]*2
        sheet1['F12'], sheet1['O12'] = [date_format(self.departure)]*2
        sheet1['B16'] = str(self.familyname).upper() + '/'
        sheet1['B19'] = str(self.firstname).upper()
        sheet1['D16'] = str(self.name).upper() + '/'
        sheet1['D19'] = str(self.lastname).upper()
        #sheet1['F16'] = date_format(self.birthday)
        #sheet1['G17'] = 'stes'#str(self.sex).upper()
        #sheet1['G16'] = self.passport
        sheet1['D21'] = str(self.goal).upper()
        #sheet1['D40'] = self.date
        #sheet1['M40'] = self.date
        sheet1['D25'] = self.placement

        sheet1['B27'],sheet1['B28'] = SplitText(str(self.hostorganization),[50,1000])
        sheet1['F23'],sheet1['B24'] = SplitText(str(self.rout),[25,1000])
        sheet1['E31'],sheet1['B32'],sheet1['B33'],sheet1['B34'] = SplitText(str(self.additionalinfo),[25,50,50,1000])

        # Внесение членов группы
        for i,g in enumerate(GroupMembers.objects.filter(form2=self)):
            for col in [0,9]:
                sheet1.cell(column=2+col,row=50+i*2).value = str(g.familyname).upper() + '/'
                sheet1.cell(column=2+col,row=51+i*2).value = str(g.name).upper()
                sheet1.cell(column=4+col,row=50+i*2).value = str(g.firstname).upper() + '/'
                sheet1.cell(column=4+col,row=51+i*2).value = str(g.lastname).upper()
                sheet1.cell(column=6+col,row=50+i*2).value = date_format(g.birthday)
                sheet1.cell(column=7+col,row=50+i*2).value = g.passport
                sheet1.cell(column=8+col,row=50+i*2).value = g.nationality


        for r in [14]+list(range(48, 70, 2)):
            for c in range(2,9):
                sheet1.cell(column=c,row = r).border = ulr
            for c in range(11, 18):
                sheet1.cell(column=c, row=r).border = ulr

        for r in [15] + list(range(49, 71, 2)):
            for c in range(2, 9):
                sheet1.cell(column=c, row=r).border = blr
            for c in range(11, 18):
                sheet1.cell(column=c, row=r).border = blr

        for r in [16,19,17,18]:
            for c in range(2,9):
                sheet1.cell(column=c, row=r).border = tlrb
            for c in range(11, 18):
                sheet1.cell(column=c, row=r).border = tlrb

        for r in [25]:
            for c in range(2,9):
                sheet1.cell(column=c, row=r).border = b
            for c in range(11, 18):
                sheet1.cell(column=c, row=r).border = b

        for r in [4,5]:
            for c in [72,74,39,41]:
                sheet1.cell(column=r, row=c).border = b

        for r in [13,14]:
            for c in [72,74,39,41]:
                sheet1.cell(column=r, row=c).border = b

        for c in range(2,18):
            sheet1.cell(column=45,row=c).border = t

        pech = openpyxl.drawing.image.Image('static/xlsx/image823.png')
        sheet1.add_image(pech, 'B35')
        pech2 = openpyxl.drawing.image.Image('static/xlsx/image823.png')
        sheet1.add_image(pech2, 'K35')
        pod = openpyxl.drawing.image.Image('static/xlsx/image853.png')
        sheet1.add_image(pod, 'D35')
        pod2 = openpyxl.drawing.image.Image('static/xlsx/image853.png')
        sheet1.add_image(pod2,'M35')

        pech = openpyxl.drawing.image.Image('static/xlsx/image823.png')
        sheet1.add_image(pech, 'B67')
        pech2 = openpyxl.drawing.image.Image('static/xlsx/image823.png')
        sheet1.add_image(pech2, 'K67')
        pod = openpyxl.drawing.image.Image('static/xlsx/image853.png')
        sheet1.add_image(pod, 'D67')
        pod2 = openpyxl.drawing.image.Image('static/xlsx/image853.png')
        sheet1.add_image(pod2,'M67')



    def FormXlsx_single(self,fin):
        self.fname = fin
        self.wb = openpyxl.Workbook()
        self.wb = openpyxl.load_workbook(filename=self.fname)
        #sheet1 = self.wb["Лист1"]
        sheet1 = self.wb.active

        if str(self.partner).lower() == 'genvisa':
            sheet1['F2'] = '130319 -                USA ' + \
            datetime.datetime.strptime(str(self.entry), '%Y-%m-%d').date().strftime("%d/%m")

        sheet1['B5'] = 'визовое приглашение № ' + str(self.invitation_number)
        sheet1['D7'] = self.multiplicity
        sheet1['D9'] = self.nationality
        sheet1['C11'], sheet1['L11'] = [date_format(self.entry)]*2
        sheet1['F11'], sheet1['O11'] = [date_format(self.departure)]*2
        sheet1['C13'] = str(self.familyname).upper() + '/' + str(self.firstname).upper()
        sheet1['E15'] = str(self.name).upper() + '/' + str(self.lastname).upper()
        sheet1['D17'] = date_format(self.birthday)
        sheet1['G17'] = str(self.sex).upper()
        sheet1['D19'] = self.passport
        sheet1['D21'] = str(self.goal).upper()
        #sheet1['D40'] = self.date
        #sheet1['M40'] = self.date
        sheet1['D25'] = self.placement
        sheet1['B27'],sheet1['B28'] = SplitText(str(self.hostorganization),[50,1000])
        sheet1['F23'],sheet1['B24'] = SplitText(str(self.rout),[25,1000])
        sheet1['E31'],sheet1['B32'],sheet1['B33'],sheet1['B34'] = SplitText(str(self.additionalinfo),[25,50,50,1000])

        pech = openpyxl.drawing.image.Image('static/xlsx/image823.png')
        pech2 = openpyxl.drawing.image.Image('static/xlsx/image823.png')
        pod = openpyxl.drawing.image.Image('static/xlsx/image853.png')
        pod2 = openpyxl.drawing.image.Image('static/xlsx/image853.png')
        sheet1.add_image(pech, 'A34')
        sheet1.add_image(pech2,'K34')
        sheet1.add_image(pod, 'C34')
        sheet1.add_image(pod2,'M34')

        for c in range(1,18):
            sheet1.cell(column=c,row=25).border = b

        sheet1['H25'].border = br
        sheet1['E41'].border = b
        sheet1['E39'].border = b
        sheet1['N41'].border = b
        sheet1['N39'].border = b
        sheet1['A25'].border = l

    def GenerateXlsx(self,fout):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename='+fout
        self.wb.save(response)
        #self.wb.save(filename='test.xlsx')
        return response

    def GeneratePdf(self,fout):
        self.wb.save('tmp1.xlsx')
        os.system('libreoffice --headless --convert-to pdf:calc_pdf_Export --outdir pdf/ tmp1.xlsx')

        with open('pdf/tmp1.pdf', 'rb') as pdf:
            response = HttpResponse(pdf.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'inline;filename=some_file.pdf'
        return response






class GroupMembers(models.Model):
    form2 = models.ForeignKey(Form2,on_delete=models.CASCADE)
    firstname = models.CharField(max_length=30,default='')
    familyname = models.CharField(max_length=30, default='')
    lastname = models.CharField(max_length=30,default='')
    name = models.CharField(max_length=30, default='')
    birthday = models.DateField(default='')
    passport = models.TextField(default='')
    nationality = models.TextField(default='')

    def __str__(self):
        return self.familyname



#https://djbook.ru/rel1.6/topics/forms/modelforms.html

