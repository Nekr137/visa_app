
import openpyxl,re

class Xlsx:
    def __init__(self,fname):
        print('XLSX init')
        self.wb = openpyxl.Workbook()
        self.fname = fname
        self.open()

    def WriteForm1(self,f1):
        sheet1 = self.wb.get_sheet_by_name("Лист1")

        sheet1['F2'] = f1.get('confirmation')
        sheet1['D5'] = f1.get('multiplicity')
        sheet1['D7'] = f1.get('nationality')
        sheet1['C9'] = f1.get('entry')
        sheet1['F9'] = f1.get('departure')
        sheet1['C11'] = f1.get('lastname')+'/'+f1.get('familyname')
        sheet1['E13'] = f1.get('firstname')+'/'+f1.get('name')
        sheet1['D15'] = self.date_format(f1.get('birthday'))
        sheet1['G15'] = f1.get('sex')
        sheet1['D17'] = f1.get('passport')
        sheet1['D19'] = f1.get('goal')
        sheet1['D40'] = f1.get('date')
        sheet1['M40'] = f1.get('date')

    def date_format(self,d):
        d = re.findall(r'\d+',d)
        d.reverse()
        return '.'.join(d)


    def open(self):
        print('XLSX OPEN')
        try:
            self.wb = openpyxl.load_workbook(filename=self.fname)
            print('FileOpened!')
            #self.ws1 = self.wb.active
            return 1
        except:
            print(self.fname,' not found!')
            return 0
            # sheet1 = self.wb.get_sheet_by_name("Sheet")
            # self.wb.remove_sheet(sheet1)
            # self.ws1 = self.wb.create_sheet("Export Products Sheet")
            # self.ws2 = self.wb.create_sheet("Export Groups Sheet")
            # up = 'Код_товара    Название_позиции	Ключевые_слова	Описание	Тип_товара	Цена	Валюта	Единица_измерения	Минимальный_объем_заказа	Оптовая_цена	Минимальный_заказ_опт	Ссылка_изображения	Наличие	Номер_группы	Название_группы	Адрес_подраздела	Возможность_поставки	Срок_поставки	Способ_упаковки	Уникальный_идентификатор	Идентификатор_товара	Идентификатор_подраздела	Идентификатор_группы	Производитель	Гарантийный_срок	Страна_производитель	Скидка	ID_группы_разновидностей'
            # up += '\tНазвание_Характеристики\tИзмерение_Характеристики\tЗначение_Характеристики' * 40
            # up = re.split(r'\s+',up)
            # self.ws1.append(up)
            # up = 'Номер_группы\tНазвание_группы\tИдентификатор_группы\tНомер_родителя\tИдентификатор_родителя'
            # up = re.split(r'\s+', up)
            # self.ws2.append(up)

    def WriteGroup(self,d,place):
        ws2 = self.wb.get_sheet_by_name("Export Groups Sheet")
        ws2[place] = d
        ws1 = self.wb.get_sheet_by_name("Export Products Sheet")

        ws1.append(product_list)

    def Save(self):
        self.wb.save(filename='test.xlsx')

class Xlsx2:
    def __init__(self,fname):
        self.wb = openpyxl.Workbook()
        self.fname = 'data.xlsx'


def open(self):
    try:
        self.wb = openpyxl.load_workbook(filename=self.fname)
    except:
        pass
    self.ws1 = self.wb.active


def WriteCell(self ,cell ,d):
    self.ws1[cell ] = d

def Close(self):
    self.wb.save(filename=self.fname)

